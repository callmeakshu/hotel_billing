import os
import sqlite3
from datetime import datetime, date, timedelta
import urllib.parse
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash, g

app = Flask(__name__)
app.secret_key = 'change-this-secret-key'
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'hotel_billing.db')

downloads_dir = os.path.join(app.static_folder, 'downloads')
os.makedirs(downloads_dir, exist_ok=True)


def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    with get_db_connection() as conn:
        conn.execute(
            '''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                hotel_name TEXT NOT NULL,
                email TEXT NOT NULL UNIQUE,
                password TEXT NOT NULL,
                contact TEXT NOT NULL
            )
            '''
        )
        conn.execute(
            '''
            CREATE TABLE IF NOT EXISTS items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                price REAL NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
            '''
        )
        conn.execute(
            '''
            CREATE TABLE IF NOT EXISTS bills (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                total_amount REAL NOT NULL,
                date_time TEXT NOT NULL,
                customer_mobile TEXT,
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
            '''
        )
        conn.execute(
            '''
            CREATE TABLE IF NOT EXISTS bill_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bill_id INTEGER NOT NULL,
                item_name TEXT NOT NULL,
                quantity INTEGER NOT NULL,
                price REAL NOT NULL,
                subtotal REAL NOT NULL,
                FOREIGN KEY(bill_id) REFERENCES bills(id)
            )
            '''
        )
        conn.commit()


init_db()


def login_required(view):
    @wraps(view)
    def wrapped_view(**kwargs):
        if g.user is None:
            return redirect(url_for('login'))
        return view(**kwargs)
    return wrapped_view


def format_datetime(dt_str):
    """Convert ISO datetime string to DD-MM-YYYY HH:MM format (IST)"""
    try:
        dt = datetime.strptime(dt_str, '%Y-%m-%d %H:%M:%S')
        return dt.strftime('%d-%m-%Y %H:%M')
    except:
        return dt_str


app.jinja_env.filters['format_dt'] = format_datetime


@app.before_request
def load_logged_in_user():
    user_id = session.get('user_id')
    if user_id is None:
        g.user = None
    else:
        conn = get_db_connection()
        g.user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        conn.close()


@app.route('/')
def home():
    if g.user:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        hotel_name = request.form['hotel_name'].strip()
        email = request.form['email'].strip().lower()
        password = request.form['password'].strip()
        contact = request.form['contact'].strip()

        if not hotel_name or not email or not password or not contact:
            flash('Please fill in all fields.', 'error')
            return render_template('register.html')

        if '@' not in email or '.' not in email:
            flash('Please enter a valid email address.', 'error')
            return render_template('register.html')

        if not contact.isdigit() or len(contact) != 10:
            flash('Contact number must be 10 digits.', 'error')
            return render_template('register.html')

        conn = get_db_connection()
        existing = conn.execute('SELECT id FROM users WHERE email = ?', (email,)).fetchone()
        if existing:
            conn.close()
            flash('Email is already registered.', 'error')
            return render_template('register.html')

        conn.execute(
            'INSERT INTO users (hotel_name, email, password, contact) VALUES (?, ?, ?, ?)',
            (hotel_name, email, password, contact)
        )
        conn.commit()
        conn.close()

        flash('Registration successful. Please log in.', 'success')
        return redirect(url_for('login'))

    return render_template('register.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email'].strip().lower()
        password = request.form['password'].strip()
        conn = get_db_connection()
        user = conn.execute('SELECT * FROM users WHERE email = ? AND password = ?', (email, password)).fetchone()
        conn.close()

        if user is None:
            flash('Invalid email or password.', 'error')
            return render_template('login.html')

        session.clear()
        session['user_id'] = user['id']
        session['hotel_name'] = user['hotel_name']
        flash(f'Welcome back, {user["hotel_name"]}!', 'success')
        return redirect(url_for('dashboard'))

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out successfully.', 'success')
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db_connection()
    total_items = conn.execute('SELECT COUNT(*) AS count FROM items WHERE user_id = ?', (g.user['id'],)).fetchone()['count']
    total_bills = conn.execute('SELECT COUNT(*) AS count FROM bills WHERE user_id = ?', (g.user['id'],)).fetchone()['count']
    total_revenue = conn.execute('SELECT IFNULL(SUM(total_amount), 0) AS sum FROM bills WHERE user_id = ?', (g.user['id'],)).fetchone()['sum']
    conn.close()
    return render_template('dashboard.html', total_items=total_items, total_bills=total_bills, total_revenue=total_revenue)


@app.route('/items', methods=['GET', 'POST'])
@login_required
def items():
    conn = get_db_connection()
    if request.method == 'POST':
        name = request.form['name'].strip()
        price = request.form['price'].strip()
        if not name or not price:
            flash('Please provide both item name and price.', 'error')
        else:
            try:
                price_value = float(price)
                conn.execute(
                    'INSERT INTO items (user_id, name, price) VALUES (?, ?, ?)',
                    (g.user['id'], name, price_value)
                )
                conn.commit()
                flash('Item added successfully.', 'success')
            except ValueError:
                flash('Price must be a valid number.', 'error')
    items_list = conn.execute('SELECT * FROM items WHERE user_id = ?', (g.user['id'],)).fetchall()
    conn.close()
    return render_template('items.html', items=items_list)


@app.route('/items/edit/<int:item_id>', methods=['GET', 'POST'])
@login_required
def edit_item(item_id):
    conn = get_db_connection()
    item = conn.execute('SELECT * FROM items WHERE id = ? AND user_id = ?', (item_id, g.user['id'])).fetchone()
    if item is None:
        conn.close()
        flash('Item not found.', 'error')
        return redirect(url_for('items'))

    if request.method == 'POST':
        name = request.form['name'].strip()
        price = request.form['price'].strip()
        if not name or not price:
            flash('Please provide item name and price.', 'error')
        else:
            try:
                price_value = float(price)
                conn.execute(
                    'UPDATE items SET name = ?, price = ? WHERE id = ? AND user_id = ?',
                    (name, price_value, item_id, g.user['id'])
                )
                conn.commit()
                flash('Item updated successfully.', 'success')
                conn.close()
                return redirect(url_for('items'))
            except ValueError:
                flash('Price must be a valid number.', 'error')
    conn.close()
    items_list = get_db_connection().execute('SELECT * FROM items WHERE user_id = ?', (g.user['id'],)).fetchall()
    return render_template('items.html', items=items_list, edit_item=item)


@app.route('/items/delete/<int:item_id>')
@login_required
def delete_item(item_id):
    conn = get_db_connection()
    conn.execute('DELETE FROM items WHERE id = ? AND user_id = ?', (item_id, g.user['id']))
    conn.commit()
    conn.close()
    flash('Item deleted successfully.', 'success')
    return redirect(url_for('items'))


@app.route('/billing', methods=['GET', 'POST'])
@login_required
def billing():
    conn = get_db_connection()
    items_list = conn.execute('SELECT * FROM items WHERE user_id = ?', (g.user['id'],)).fetchall()
    bill_summary = None
    if request.method == 'POST':
        action = request.form.get('action', 'preview')
        customer_mobile = request.form.get('customer_mobile', '').strip()

        if action == 'preview':
            item_ids = request.form.getlist('item_id')
            quantities = request.form.getlist('quantity')
            if customer_mobile and (not customer_mobile.isdigit() or len(customer_mobile) != 10):
                flash('Customer mobile must be 10 digits.', 'error')
                conn.close()
                return render_template('billing.html', items=items_list)

            bill_items = []
            total_amount = 0.0
            for item_id, qty in zip(item_ids, quantities):
                if not item_id or not qty:
                    continue
                try:
                    quantity = int(qty)
                    if quantity <= 0:
                        continue
                except ValueError:
                    continue

                item = conn.execute(
                    'SELECT name, price FROM items WHERE id = ? AND user_id = ?', (item_id, g.user['id'])
                ).fetchone()
                if item is None:
                    continue

                subtotal = item['price'] * quantity
                bill_items.append({
                    'item_name': item['name'],
                    'quantity': quantity,
                    'price': item['price'],
                    'subtotal': subtotal,
                })
                total_amount += subtotal

            if not bill_items:
                flash('Please add at least one valid item to the bill.', 'error')
                conn.close()
                return render_template('billing.html', items=items_list)

            bill_summary = {
                'hotel_name': g.user['hotel_name'],
                'date_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'bill_items': bill_items,
                'total_amount': total_amount,
                'customer_mobile': customer_mobile,
            }
        elif action == 'send':
            item_ids = request.form.getlist('item_id')
            quantities = request.form.getlist('quantity')
            if not customer_mobile or not customer_mobile.isdigit() or len(customer_mobile) != 10:
                flash('Customer mobile must be provided and be 10 digits to send via WhatsApp.', 'error')
                conn.close()
                return render_template('billing.html', items=items_list)

            bill_items = []
            total_amount = 0.0
            for item_id, qty in zip(item_ids, quantities):
                if not item_id or not qty:
                    continue
                try:
                    quantity = int(qty)
                    if quantity <= 0:
                        continue
                except ValueError:
                    continue

                item = conn.execute(
                    'SELECT name, price FROM items WHERE id = ? AND user_id = ?',
                    (item_id, g.user['id'])
                ).fetchone()
                if item is None:
                    continue

                subtotal = item['price'] * quantity
                bill_items.append({
                    'item_name': item['name'],
                    'quantity': quantity,
                    'price': item['price'],
                    'subtotal': subtotal,
                })
                total_amount += subtotal

            if not bill_items:
                flash('Unable to save bill. Please try again.', 'error')
                conn.close()
                return render_template('billing.html', items=items_list)

            date_time = (datetime.utcnow() + timedelta(hours=5, minutes=30)).strftime('%Y-%m-%d %H:%M:%S')
            cursor = conn.cursor()
            cursor.execute(
                'INSERT INTO bills (user_id, total_amount, date_time, customer_mobile) VALUES (?, ?, ?, ?)',
                (g.user['id'], total_amount, date_time, customer_mobile)
            )
            bill_id = cursor.lastrowid
            for row in bill_items:
                cursor.execute(
                    'INSERT INTO bill_items (bill_id, item_name, quantity, price, subtotal) VALUES (?, ?, ?, ?, ?)',
                    (bill_id, row['item_name'], row['quantity'], row['price'], row['subtotal'])
                )
            conn.commit()
            conn.close()

            # Generate WhatsApp message with full bill details
            bill_text = f"Thank you for visiting {g.user['hotel_name']}!\n\nBill Details:\n"
            for item in bill_items:
                bill_text += f"{item['item_name']}: {item['quantity']} x ₹{item['price']:.2f} = ₹{item['subtotal']:.2f}\n"
            bill_text += f"\nTotal Amount: ₹{total_amount:.2f}\nDate & Time: {format_datetime(date_time)}\n\nWe look forward to seeing you again!"
            whatsapp_url = generate_whatsapp_link(customer_mobile, bill_text)
            return redirect(whatsapp_url)
    conn.close()
    return render_template('billing.html', items=items_list, bill_summary=bill_summary)


def normalize_phone(phone):
    phone = phone.strip().replace(' ', '').replace('-', '')
    if phone.startswith('+'):
        phone = phone[1:]
    if phone.startswith('0'):
        phone = phone[1:]
    if len(phone) == 10:
        phone = '91' + phone
    return phone


def generate_whatsapp_link(phone, message):
    phone = normalize_phone(phone)
    encoded = urllib.parse.quote(message)
    return f"https://wa.me/{phone}?text={encoded}"


@app.route('/revenue', methods=['GET', 'POST'])
@login_required
def revenue():
    conn = get_db_connection()
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    filter_mode = request.args.get('mode', 'today')

    if filter_mode == 'range' and start_date and end_date:
        query = '''
            SELECT * FROM bills
            WHERE user_id = ? AND date(date_time) BETWEEN ? AND ?
            ORDER BY date_time DESC
        '''
        bills = conn.execute(query, (g.user['id'], start_date, end_date)).fetchall()
        summary = conn.execute(
            'SELECT IFNULL(SUM(total_amount), 0) AS total FROM bills WHERE user_id = ? AND date(date_time) BETWEEN ? AND ?',
            (g.user['id'], start_date, end_date)
        ).fetchone()
    else:
        today = date.today().isoformat()
        query = 'SELECT * FROM bills WHERE user_id = ? AND date(date_time) = ? ORDER BY date_time DESC'
        bills = conn.execute(query, (g.user['id'], today)).fetchall()
        summary = conn.execute(
            'SELECT IFNULL(SUM(total_amount), 0) AS total FROM bills WHERE user_id = ? AND date(date_time) = ?',
            (g.user['id'], today)
        ).fetchone()
        start_date = today
        end_date = today
        filter_mode = 'today'

    total_revenue = summary['total']
    total_bills = len(bills)
    conn.close()
    return render_template('revenue.html', bills=bills, total_revenue=total_revenue,
                           total_bills=total_bills, start_date=start_date, end_date=end_date,
                           filter_mode=filter_mode)


@app.route('/revenue/delete', methods=['POST'])
@login_required
def delete_history():
    email = request.form.get('email', '').strip().lower()
    password = request.form.get('password', '').strip()
    if email != g.user['email'] or password != g.user['password']:
        flash('Email or password is incorrect. Cannot delete billing history.', 'error')
        return redirect(url_for('revenue'))

    conn = get_db_connection()
    bill_ids = [row['id'] for row in conn.execute('SELECT id FROM bills WHERE user_id = ?', (g.user['id'],)).fetchall()]
    if bill_ids:
        placeholders = ','.join('?' * len(bill_ids))
        conn.execute(f'DELETE FROM bill_items WHERE bill_id IN ({placeholders})', bill_ids)
    conn.execute('DELETE FROM bills WHERE user_id = ?', (g.user['id'],))
    conn.commit()
    conn.close()
    flash('Billing history deleted successfully. Revenue updated.', 'success')
    return redirect(url_for('revenue'))


@app.route('/download-excel')
@login_required
def download_excel():
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    filter_mode = request.args.get('mode', 'today')

    conn = get_db_connection()
    if filter_mode == 'range' and start_date and end_date:
        query = '''
            SELECT * FROM bills
            WHERE user_id = ? AND date(date_time) BETWEEN ? AND ?
            ORDER BY date_time DESC
        '''
        bills = conn.execute(query, (g.user['id'], start_date, end_date)).fetchall()
    else:
        today = date.today().isoformat()
        query = 'SELECT * FROM bills WHERE user_id = ? AND date(date_time) = ? ORDER BY date_time DESC'
        bills = conn.execute(query, (g.user['id'], today)).fetchall()
    conn.close()

    try:
        from openpyxl import Workbook
    except ImportError:
        flash('Excel generation not available. Please install openpyxl.', 'error')
        return redirect(url_for('revenue'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Bill Records"

    # Headers
    ws['A1'] = 'Date & Time'
    ws['B1'] = 'Total Amount'
    ws['C1'] = 'Customer Mobile'

    # Data
    for i, bill in enumerate(bills, start=2):
        ws[f'A{i}'] = format_datetime(bill['date_time'])
        ws[f'B{i}'] = f"Rs. {bill['total_amount']:.2f}"
        ws[f'C{i}'] = bill['customer_mobile'] or 'N/A'

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='bill_records.xlsx')


@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    if request.method == 'POST':
        hotel_name = request.form['hotel_name'].strip()
        email = request.form['email'].strip().lower()
        contact = request.form['contact'].strip()
        new_password = request.form['password'].strip()
        current_email = request.form['current_email'].strip().lower()
        current_password = request.form['current_password'].strip()

        if not hotel_name or not email or not contact or not current_email or not current_password:
            flash('Please fill all fields and provide current email and password.', 'error')
            return render_template('settings.html')

        if current_email != g.user['email'] or current_password != g.user['password']:
            flash('Current email or password is incorrect.', 'error')
            return render_template('settings.html')

        if not contact.isdigit() or len(contact) != 10:
            flash('Contact number must be 10 digits.', 'error')
            return render_template('settings.html')

        conn = get_db_connection()
        if email != g.user['email']:
            existing = conn.execute('SELECT id FROM users WHERE email = ? AND id != ?', (email, g.user['id'])).fetchone()
            if existing:
                conn.close()
                flash('This email is already registered.', 'error')
                return render_template('settings.html')

        if new_password:
            conn.execute(
                'UPDATE users SET hotel_name = ?, email = ?, contact = ?, password = ? WHERE id = ?',
                (hotel_name, email, contact, new_password, g.user['id'])
            )
        else:
            conn.execute(
                'UPDATE users SET hotel_name = ?, email = ?, contact = ? WHERE id = ?',
                (hotel_name, email, contact, g.user['id'])
            )
        conn.commit()
        conn.close()

        session['hotel_name'] = hotel_name
        flash('Profile settings updated successfully.', 'success')
        return redirect(url_for('settings'))
    return render_template('settings.html')


@app.route('/help')
@login_required
def help():
    return render_template('help.html')


if __name__ == '__main__':
    app.run(debug=True)
